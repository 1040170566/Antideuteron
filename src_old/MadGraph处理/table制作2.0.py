"""
更新代码逻辑，输出excel但直接读取数据
时间：2024.1
作者：王骁
"""

# import madgraph数据提取 as mg
import openpyxl as op
import pandas as pd
import numpy as np
import sys
# import os

# 单元格格式设置
from openpyxl.styles import Alignment
alignment = Alignment(horizontal='center', vertical='center')

############################################
# 参数获取
M_DM = int(sys.argv[1])  # 暗物质质量
DATA_path = sys.argv[2]  # 数据文件路径

# 数据加载
DATA = pd.read_csv(DATA_path, header=0, delimiter='\\s+')
bins_x = DATA['x=T/M_DM'].values  # 获取刻度，x=T/M
bins_lgx = np.log10(bins_x)  # bins化为lg(x)刻度
numbers_xT_antideuteron = DATA['dN/dlgx'].values  # 获取数据，dN/dlgx

############################################

############################################
# 输出区
# 输出至excel制作table
# 以列为单位，依次表示暗物质质量M_DM，lg(x)，dN/dlg(x)。

# 读取excel
output_path = r'处理结果/数据/AtProduction_antideuterons.xlsx'
wb = op.load_workbook(filename=output_path)
ws = wb.worksheets[0]
ws.title = '方法一'
wb.copy_worksheet(ws)  # sheet备份

length_bins = len(bins_x)
# 选择开始插入行的位置
col_A = tuple(ws.iter_cols(min_row=2, max_col=1))[0]
if not col_A or int(col_A[-1].value) < M_DM:
    row_insert = ws.max_row + 1
else:
    row_insert = 2
    while True:
        if M_DM < np.ceil(col_A[row_insert - 2].value):
            ws.insert_rows(row_insert, length_bins)
            break
        if M_DM == np.ceil(col_A[row_insert - 2].value):
            break
        if M_DM > np.ceil(col_A[row_insert - 2].value):
            row_insert += length_bins

# 选择插入的列指标
print('产生通道有：')
row_1 = tuple(ws.iter_rows(min_col=3, max_col=ws.max_column-1, max_row=1))[0]
for n in range(len(row_1)):
    print(f'{n+1}.{row_1[n].value}')
column_insert = int(input('你的选择：'))+2
# column_insert = 3
# print('你的选择：', column_insert)


# 向第一张工作表中输出数据
for i in range(length_bins):
    # 写入
    ws.cell(row=row_insert, column=1, value=M_DM).alignment = alignment
    ws.cell(row=row_insert, column=2, value=bins_lgx[i]).alignment = alignment
    ws.cell(row=row_insert, column=column_insert, value=numbers_xT_antideuteron[i]).alignment = alignment
    # ws.cell(row=row_insert, column=ws.max_column, value=mg.EventsNumber).alignment = alignment

    row_insert += 1

# 保存并打开
wb.save(output_path)
# os.startfile(r"/home/wangxiao/文档/dnde.xlsx") # 适用于windows
# os.system('libreoffice "/home/wangxiao/文档/dnde.xlsx"') # 适用于Linux
###########################################################
