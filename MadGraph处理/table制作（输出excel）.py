# import madgraph数据提取 as mg
import openpyxl as op
import pandas as pd
import numpy as np
import sys
# import os
"""
##############################################
# 数据处理区
# 核子动能集
T_nu = [nu.four_momentum[-1] - nu.mass for nu in mg.antinucleon]
Tmax_nu, Tmin_nu = max(T_nu), min(T_nu)

bins_T_nu = np.linspace(Tmin_nu, Tmax_nu, mg.bins_number + 1)
# bins_T_nu = nplog(Tmin_nu, Tmax_nu, bins_number + 1)

# 反质子动能归一化分布
T_pbar = [(p.four_momentum[-1] - p.mass) for p in mg.pbar]
numbers_T_pbar = np.histogram(T_pbar, bins=bins_T_nu)[0].astype(float)
nor_dndT_pbar = numbers_T_pbar / np.diff(bins_T_nu) / mg.EventsNumber
# for x in range(len(numbers_T_pbar)):
#     numbers_T_pbar[x] /= (bins_T_nu[x + 1] - bins_T_nu[x]) * mg.EventsNumber

# 反中子动能归一化分布
T_nbar = [(n.four_momentum[-1] - n.mass) for n in mg.nbar]
numbers_T_nbar = np.histogram(T_nbar, bins=bins_T_nu)[0].astype(float)
nor_dndT_nbar = numbers_T_nbar / np.diff(bins_T_nu) / mg.EventsNumber
# for x in range(len(numbers_T_nbar)):
#     numbers_T_nbar[x] /= (bins_T_nu[x + 1] - bins_T_nu[x]) * mg.EventsNumber

# 聚结参数 * 反氘动量
B_AP_A = mg.M_antideuteron / mg.mass_pbar / mg.mass_nbar * mg.Pcoal ** 3 / 6

p_nu = np.sqrt((bins_T_nu + mg.mass_pbar) ** 2 - mg.mass_pbar ** 2)  # 每核子动量
bins_P_antideu = p_nu*2  # 反氘动量bins
nor_dndT_antideuteron = B_AP_A / bins_P_antideu[:-1] * nor_dndT_pbar * nor_dndT_nbar  # 反氘动能归一化分布

bins_T_antideu = np.sqrt(bins_P_antideu ** 2 + mg.M_antideuteron ** 2) - mg.M_antideuteron  # 反氘动能bins

# bins_xT_antideu = [t / M_DM for t in bins_T_antideu]  # bins化为x刻度，x=T/M
bins_lgx_antideu = np.log10(bins_T_antideu / mg.M_DM)  # bins化为lg(x)刻度，x=T/M

numbers_xT_antideuteron = nor_dndT_antideuteron * bins_T_antideu[:-1] * np.log(10)
############################################
"""
# TODO: 优化计算速度。暂时修改到这里，下次再改。现在感觉这种统计方法不太好，应该直接用pppc4的bins来统计。

############################################
# 参数获取
M_DM = int(sys.argv[1])  # 暗物质质量
DATA_path = sys.argv[2]  # 数据文件路径

# 数据加载
DATA = pd.read_csv(DATA_path, header=0, delimiter='\\s+')
bins_x = DATA['x=T/M_DM'].values  # bins化为lg(x)刻度，x=T/M
bins_lgx_antideu = np.log10(bins_x)  # bins化为lg(x)刻度，x=T/M

############################################

############################################
# 输出区
# 输出至excel制作table
# 以列为单位，依次表示暗物质质量M_DM，lg(x)，dN/dlg(x)。

wb = op.load_workbook(filename=r'处理结果/数据/AtProduction_antideuterons.xlsx')
ws = wb.worksheets[0]
ws.title = '方法一'
wb.copy_worksheet(ws)  # sheet备份

length_bins = len(bins_x)
# 选择开始插入行的位置
col_A = tuple(ws.iter_cols(min_row=2, max_col=1))[0]
if not col_A or int(col_A[-1].value) < M_DM:
    row_insert = ws.max_row + 1
else:
    row_insert = 2
    while True:
        if M_DM < int(col_A[row_insert - 2].value):
            ws.insert_rows(row_insert, length_bins)
            break
        if M_DM == int(col_A[row_insert - 2].value):
            break
        if M_DM > int(col_A[row_insert - 2].value):
            row_insert += length_bins

# 选择插入的列指标
print('产生通道有：')
row_1 = tuple(ws.iter_rows(min_col=3, max_col=ws.max_column-1, max_row=1))[0]
for n in range(len(row_1)):
    print(f'{n+1}.{row_1[n].value}')
column_insert = int(input('你的选择：'))+2
# column_insert = 3
# print('你的选择：', column_insert)

# 单元格格式设置
from openpyxl.styles import Alignment
alignment = Alignment(horizontal='center', vertical='center')

# 向第一张工作表中输出数据
for lgx in bins_lgx:
    # 查找lgx在bins上的位置
    if lgx < bins_lgx_antideu[0]:
        index = 0
        # value = numbers_xT_antideuteron[0]
    elif lgx >= bins_lgx_antideu[-1]:
        index = -1
        # value = numbers_xT_antideuteron[-1]
    else:
        for index, value in enumerate(bins_lgx_antideu):
            if value <= lgx < bins_lgx_antideu[index + 1]:
                break
    # 写入
    ws.cell(row=row_insert, column=1, value=M_DM).alignment = alignment
    ws.cell(row=row_insert, column=2, value=lgx).alignment = alignment
    ws.cell(row=row_insert, column=column_insert, value=numbers_xT_antideuteron[index]).alignment = alignment
    ws.cell(row=row_insert, column=ws.max_column, value=mg.EventsNumber).alignment = alignment

    row_insert += 1

# 保存并打开
wb.save(r"/home/wangxiao/document/data_output/AtProduction_antideuterons.xlsx")
# os.startfile(r"/home/wangxiao/文档/dnde.xlsx") # 适用于windows
# os.system('libreoffice "/home/wangxiao/文档/dnde.xlsx"') # 适用于Linux
###########################################################
